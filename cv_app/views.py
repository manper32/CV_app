from django.shortcuts import render
from django.http import HttpResponse
from datetime import datetime
import psycopg2
import math
import pandas as pd
from openpyxl import Workbook
import csv

def psql_pdc(query):
    #credenciales PostgreSQL produccion
    connP_P = {
    	'host' : '10.150.1.74',
    	'port' : '5432',
    	'user':'postgres',
    	'password':'cobrando.bi.2020',
    	'database' : 'postgres'}

    #conexion a PostgreSQL produccion
    conexionP_P = psycopg2.connect(**connP_P)
    #print('\nConexión con el servidor PostgreSQL produccion establecida!')
    cursorP_P = conexionP_P.cursor ()
    #ejecucion query telefonos PostgreSQL
    cursorP_P.execute(query)
    anwr = cursorP_P.fetchall()
    
    cursorP_P.close()
    conexionP_P.close()
    return anwr
    
def to_horiz(anwr_P,name):
    #vertical horizontal telefonos
    anwr_P1 = anwr_P.pivot(index=0,columns=1)
    anwr_P1["deudor_id"] = anwr_P1.index
    
    #nombre de telefonos
    col1 = []
    i=0
    for i in range(anwr_P1.shape[1]-1):
        col1.append(name+str(i+1))
    col1.append("deudor_id")
    
    anwr_P1.columns = col1
    
    return anwr_P1

def csv_o(fn,name):
    response = HttpResponse(content_type = "text/csv")
    content = "attachment; filename = %s"%name
    response["Content-Disposition"] = content

    fn2 = [tuple(x) for x in fn.values]
    writer = csv.writer(response,delimiter ='|')
    writer.writerow(fn.columns)
    writer.writerows(fn2)

    return response

def excel(fn,name):
    wb = Workbook()
    ws = wb.active

    k = 0
    a = pd.DataFrame(fn.columns)

    for k in range(a.shape[0]):
        ws.cell(row = 1, column = k+1).value = a.iloc[k,0]

    i=0
    j=0

    for i in range(fn.shape[0]):
        for j in range(0,fn.shape[1]):
            try:
                ws.cell(row = i+2, column = j+1).value = fn.iloc[i,j]
            except:
                pass
            
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = %s"%name
    response["Content-Disposition"] = content
    wb.save(response)
    return response

def excel_CV_COL(request):
    today = datetime.now()
    tablename = "CV_Col"+today.strftime("%Y%m%d%H")+'.xlsx'

    with open("/home/manuel/Documentos/Django/cv_app_d/cv_app_d/Plantillas/Colp/QueryTel_COL.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("/home/manuel/Documentos/Django/cv_app_d/cv_app_d/Plantillas/Colp/QueryCor_COL.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("/home/manuel/Documentos/Django/cv_app_d/cv_app_d/Plantillas/Colp/QueryDir_COL.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("/home/manuel/Documentos/Django/cv_app_d/cv_app_d/Plantillas/Colp/QueryCV_COL.txt","r") as f4:
        queryP_cons = f4.read()

    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone')
    infC = to_horiz(anwr_C,'mail')
    infD = to_horiz(anwr_D,'address')


    df = df.rename(columns={0:'rownumber',
                            1:'obligacion_id',
                            2:'deudor_id',
                            3:'unico',
                            4:'estado',
                            5:'tipo_cliente',
                            6:'nombre',
                            7:'producto',
                            8:'initial_bucket',
                            9:'ciudad',
                            10:'sucursal',
                            11:'tipo_prod',
                            12:'dias_mora_inicial',
                            13:'dias_mora_actual',
                            14:'rango_mora_inicial',
                            15:'rango_mora_final',
                            16:'rango',
                            17:'suma_pareto',
                            18:'rango_pareto',
                            19:'fcast',
                            20:'fdesem',
                            21:'vrdesem',
                            22:'saldo_total_inicial',
                            23:'saldo_total_actual',
                            24:'saldo_capital_inicial',
                            25:'saldo_capital_actual',
                            26:'saldo_vencido_inicial',
                            27:'saldo_vencido_actual',
                            28:'pagomin',
                            29:'fultpago',
                            30:'vrultpago',
                            31:'agencia',
                            32:'tasainter',
                            33:'feultref',
                            34:'ultcond',
                            35:'fasigna',
                            36:'eqasicampana',
                            37:'diferencia_pago',
                            38:'pago_preliminar',
                            39:'pago_cliente',
                            40:'min',
                            41:'tarifa',
                            42:'honorarios',
                            43:'perfil_mes_4',
                            44:'perfil_mes_3',
                            45:'perfil_mes_2',
                            46:'perfil_mes_1',
                            47:'fecha_primer_gestion',
                            48:'fecha_ultima_gestion',
                            49:'perfil_mes_actual',
                            50:'contactabilidad',
                            51:'ultimo_alo',
                            52:'descod1',
                            53:'descod2',
                            54:'asesor',
                            55:'fecha_gestion',
                            56:'telefono_mejor_gestion',
                            57:'mejorgestionhoy',
                            58:'repeticion',
                            59:'llamadas',
                            60:'sms',
                            61:'correos',
                            62:'gescall',
                            63:'visitas',
                            64:'whatsapp',
                            65:'no_contacto',
                            66:'total_gestiones',
                            67:'telefono_positivo',
                            68:'marcaciones_telefono_positivo',
                            69:'ultima_marcacion_telefono_positivo',
                            70:'fec_creacion_ult_compromiso',
                            71:'fec_pactada_ult_compromiso',
                            72:'valor_acordado_ult_compromiso',
                            73:'asesor_ult_compromiso',
                            74:'cantidad_acuerdos_mes',
                            75:'estado_acuerdo'
                            })

    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return excel(fn,tablename)


def csv_CV_Claro(request):
    today = datetime.now()
    tablename = "CV_Claro"+today.strftime("%Y%m%d%H")+'.csv'

    with open("/home/manuel/Documentos/Django/cv_app_d/cv_app_d/Plantillas/Claro/QueryTel_Claro.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("/home/manuel/Documentos/Django/cv_app_d/cv_app_d/Plantillas/Claro/QueryCor_Claro.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("/home/manuel/Documentos/Django/cv_app_d/cv_app_d/Plantillas/Claro/QueryCV_Claro.txt","r") as f4:
        queryP_cons = f4.read()
        
    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    yanwr = psql_pdc(queryP_cons)    

    #dataframes
    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    df = pd.DataFrame(yanwr)

    anwr_P1 = to_horiz(anwr_P,'phone')

    #renombrar campos correos
    anwr_C = anwr_C.rename(columns={
                            
                            0:'deudor_id',
                            1:'mail0',
                            2:'mail1'})

    #renombrar campos CV
    df = df.rename(columns={0:'rownumber',
                            1:'deudor_id',
                            2:'obligacion_id',
                            3:'nombredelcliente',
                            4:'estado',
                            5:'tipo_cliente',
                            6:'unico',
                            7:'crmorigen',
                            8:'potencialmark',
                            9:'prepotencialmark',
                            10:'writeoffmark',
                            11:'segmento_bpo',
                            12:'valorscoring',
                            13:'numeroreferenciadepago',
                            14:'monto_inicial',
                            15:'monto_ini_cuenta',
                            16:'porcentaje_descuento',
                            17:'val_descuento',
                            18:'val_a_pagar',
                            19:'deuda_real',
                            20:'val_pago',
                            21:'maxfec_pago',
                            22:'mfecha_compromiso',
                            23:'fecha_pago',
                            24:'valor_compromiso',
                            25:'estado_acuerdo',
                            26:'indicador_m4',
                            27:'indicador_m3',
                            28:'indicador_m2',
                            29:'indicador_m1',
                            30:'fecha_primer_gestion',
                            31:'fecha_ultima_gestion',
                            32:'indicador_mes_actual',
                            33:'tel_mes_actual',
                            34:'asesor_mes_actual',
                            35:'fec_indicador_mes_actual',
                            36:'contact',
                            37:'indicador_hoy',
                            38:'repeticion_mes_actual',
                            39:'llamadas_mes_actual',
                            40:'sms_mes_actual',
                            41:'correos_mes_actual',
                            42:'gescall_mes_actual',
                            43:'whatsapp_mes_actual',
                            44:'visitas',
                            45:'no_contacto_mes_actual',
                            46:'total_ges_mes_actual',
                            47:'tel_positivo',
                            48:'ult_fec_tel_pos'})

    i=0
    lin = ['no_contacto_mes_actual','gescall_mes_actual','tel_mes_actual','tel_positivo']
    for i in lin:
        df[i].fillna(0,inplace=True)
        df[i] = df[i].apply(lambda x: round(x))
        df[i] = df[i].astype('str')

    fn = pd.merge(df,anwr_P1,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,anwr_C,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return csv_o(fn,tablename)
